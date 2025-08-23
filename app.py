import streamlit as st
import pandas as pd
import numpy as np
import re
import datetime
import matplotlib.pyplot as plt
from io import BytesIO
import xlsxwriter
from openpyxl import Workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# --- User credentials (for demonstration) ---
USERNAME = "Pushpal@2025"
PASSWORD = "Pushpal@202512345"

# --- Authentication Function ---
def check_password():
    """Returns `True` if the user is authenticated."""
    if st.session_state.get('logged_in', False):
        return True

    with st.form("login"):
        st.subheader("🔐 Login Required")
        username = st.text_input("Username")
        password = st.text_input("Password", type="password")
        login = st.form_submit_button("Login")

    if login:
        if username == USERNAME and password == PASSWORD:
            st.session_state.logged_in = True
            st.rerun()
        else:
            st.error("Incorrect credentials")
    return False

# --- Helper Functions ---
def clean_level(x):
    """Extracts the integer level number from a string."""
    try:
        return int(re.search(r"(\d+)", str(x)).group(1))
    except (AttributeError, ValueError):
        return None

def generate_excel_report_dp1(df_summary, df_progression, fig1, fig2):
    """Generates an Excel report for the DP1GAME METRIX tool."""
    output = BytesIO()
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        df_summary.to_excel(writer, index=False, sheet_name='Report', startrow=0, startcol=0)
        df_progression.to_excel(writer, index=False, sheet_name='Report', startrow=0, startcol=3)
        workbook = writer.book
        worksheet = writer.sheets['Report']
        img1 = BytesIO()
        fig1.savefig(img1, format='png', bbox_inches='tight')
        img1.seek(0)
        worksheet.insert_image('G2', 'chart1.png', {'image_data': img1})
        img2 = BytesIO()
        fig2.savefig(img2, format='png', bbox_inches='tight')
        img2.seek(0)
        worksheet.insert_image('G37', 'chart2.png', {'image_data': img2})
    output.seek(0)
    return output

def generate_excel_report_gp(df_export, retention_fig, drop_fig, drop_comb_fig, version, date_selected):
    """Generates an Excel report for the Game Progression tool with color formatting."""
    output = BytesIO()
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        df_export.to_excel(writer, index=False, sheet_name='Game Progression')
        workbook = writer.book
        worksheet = writer.sheets['Game Progression']
        
        # Add title and metadata
        title = f"Game Progression Report"
        worksheet.write('M1', title, workbook.add_format({'bold': True, 'font_size': 14}))
        worksheet.write('M2', f"Version: {version}", workbook.add_format({'font_size': 10}))
        worksheet.write('M3', f"Date: {date_selected.strftime('%d-%m-%Y')}", workbook.add_format({'font_size': 10}))

        # Cell formats for drop columns
        format_dark_red = workbook.add_format({'bg_color': '#8B0000', 'font_color': 'white', 'align': 'center', 'valign': 'vcenter', 'bold': True, 'num_format': '0.00'})
        format_red = workbook.add_format({'bg_color': '#CD5C5C', 'font_color': 'white', 'align': 'center', 'valign': 'vcenter', 'bold': True, 'num_format': '0.00'})
        format_pink = workbook.add_format({'bg_color': '#FFC0CB', 'font_color': 'black', 'align': 'center', 'valign': 'vcenter', 'bold': True, 'num_format': '0.00'})
        
        # Apply conditional formatting
        for col_name in ['Game Play Drop', 'Popup Drop', 'Total Level Drop']:
            if col_name in df_export.columns:
                col_idx = df_export.columns.get_loc(col_name)
                for row_idx, value in enumerate(df_export[col_name]):
                    if pd.notna(value):
                        if value >= 10:
                            worksheet.write(row_idx + 1, col_idx, value, format_dark_red)
                        elif value >= 5:
                            worksheet.write(row_idx + 1, col_idx, value, format_red)
                        elif value >= 3:
                            worksheet.write(row_idx + 1, col_idx, value, format_pink)
                        else:
                            worksheet.write(row_idx + 1, col_idx, value)

        # Insert charts
        chart_start_row = 2
        
        img1 = BytesIO()
        retention_fig.savefig(img1, format='png', bbox_inches='tight')
        img1.seek(0)
        worksheet.insert_image(f'M{chart_start_row}', 'retention_chart.png', {'image_data': img1})

        img2 = BytesIO()
        drop_fig.savefig(img2, format='png', bbox_inches='tight')
        img2.seek(0)
        chart_start_row += 30
        worksheet.insert_image(f'M{chart_start_row}', 'drop_chart.png', {'image_data': img2})
        
        img3 = BytesIO()
        drop_comb_fig.savefig(img3, format='png', bbox_inches='tight')
        img3.seek(0)
        chart_start_row += 30
        worksheet.insert_image(f'M{chart_start_row}', 'combo_drop_chart.png', {'image_data': img3})

    output.seek(0)
    return output

def create_retention_chart(df, version, date_selected, title):
    """Creates a retention chart from a DataFrame."""
    fig, ax = plt.subplots(figsize=(15, 7))
    ax.plot(df['LEVEL_CLEAN'], df['Retention %'], linestyle='-', color='#F57C00', linewidth=2, label='RETENTION')
    ax.set_xlim(1, 100)
    ax.set_ylim(0, 120)
    ax.set_xticks(np.arange(1, 101, 1))
    ax.set_yticks(np.arange(0, 121, 10))
    ax.set_xlabel("Level", labelpad=15)
    ax.set_ylabel("% Of Users", labelpad=15)
    ax.set_title(f"{title} | Version {version} | Date: {date_selected.strftime('%d-%m-%Y')}", fontsize=12, fontweight='bold')
    xtick_labels = [f"$\\bf{{{val}}}$" if val % 5 == 0 else str(val) for val in np.arange(1, 101, 1)]
    ax.set_xticklabels(xtick_labels, fontsize=6)
    ax.tick_params(axis='x', labelsize=6)
    ax.grid(True, linestyle='--', linewidth=0.5)
    for x, y in zip(df['LEVEL_CLEAN'], df['Retention %']):
        if not np.isnan(y):
            ax.text(x, -5, f"{int(y)}", ha='center', va='top', fontsize=7)
    ax.legend(loc='lower left', fontsize=8)
    plt.tight_layout(rect=[0, 0.03, 1, 0.97])
    return fig

def create_drop_chart(df, version, date_selected, title):
    """Creates a drop chart from a DataFrame."""
    fig, ax = plt.subplots(figsize=(15, 6))
    bars = ax.bar(df['Level'], df['Drop'], color='#EF5350', label='DROP RATE')
    ax.set_xlim(1, 100)
    ax.set_ylim(0, max(df['Drop'].max(), 10) + 10)
    ax.set_xticks(np.arange(1, 101, 1))
    ax.set_yticks(np.arange(0, max(df['Drop'].max(), 10) + 11, 5))
    ax.set_xlabel("Level")
    ax.set_ylabel("% Of Users Dropped")
    ax.set_title(f"{title} | Version {version} | Date: {date_selected.strftime('%d-%m-%Y')}", fontsize=12, fontweight='bold')
    xtick_labels = [f"$\\bf{{{val}}}$" if val % 5 == 0 else str(val) for val in np.arange(1, 101, 1)]
    ax.set_xticklabels(xtick_labels, fontsize=6)
    ax.tick_params(axis='x', labelsize=6)
    ax.grid(True, linestyle='--', linewidth=0.5)
    for bar in bars:
        x = bar.get_x() + bar.get_width() / 2
        y = bar.get_height()
        ax.text(x, -2, f"{y:.0f}", ha='center', va='top', fontsize=7)
    ax.legend(loc='upper right', fontsize=8)
    plt.tight_layout()
    return fig

# --- Tool 1: DP1GAME METRIX ---
def dp1game_metrix_tool():
    st.header("📊 DP1GAME METRIX Dashboard")
    st.markdown("This tool processes retention and ad event data to provide a comprehensive game metrics report.")
    
    col1, col2 = st.columns(2)
    with col1:
        file1 = st.file_uploader("📥 Upload Retention Base File (.csv)", type=["csv"], key='file1')
    with col2:
        file2 = st.file_uploader("📥 Upload Ad Event File (.csv)", type=["csv"], key='file2')

    if not file1 and not file2:
        st.subheader("💡 Sample Input File Format")
        st.markdown("Please use the following format for your CSV files:")
        st.markdown("---")
        st.markdown("#### **Retention Base File**")
        st.image("https://i.imgur.com/k91w4sI.png", caption="Sample Retention Base Data", use_column_width=True)
        st.markdown("---")
        st.markdown("#### **Ad Event File**")
        st.image("https://i.imgur.com/8Qj9g7c.png", caption="Sample Ad Event Data", use_column_width=True)
        st.markdown("---")
    
    st.subheader("📝 Editable Fields")
    col3, col4, col5 = st.columns(3)
    with col3:
        version = st.text_input("Enter Version (e.g. v1.2.3)", value="v1.0.0", key='version1')
    with col4:
        date_selected = st.date_input("Date Selected", value=datetime.date.today(), key='date1')
    with col5:
        check_date = st.date_input("Check Date", value=datetime.date.today() + datetime.timedelta(days=1), key='check_date1')

    user_install_count = st.number_input("🔢 Optional: Enter User Install Count", min_value=0, value=None, step=1, key='install_count1')
    
    if file1 and file2:
        st.markdown("---")
        st.subheader("✅ Data Processing & Analytics")
        df1 = pd.read_csv(file1)
        df1.columns = df1.columns.str.strip().str.upper()
        level_col = next((col for col in df1.columns if col in ['LEVEL','Level', 'LEVELPLAYED', 'TOTALLEVELSPLAYED', 'LEVEL_NUMBER']), None)
        if level_col and 'USERS' in df1.columns:
            df1 = df1[[level_col, 'USERS']]
            df1['LEVEL_CLEAN'] = df1[level_col].apply(clean_level)
            df1.dropna(inplace=True)
            df1['LEVEL_CLEAN'] = df1['LEVEL_CLEAN'].astype(int)
            df1.sort_values('LEVEL_CLEAN', inplace=True)
            if user_install_count is not None and user_install_count > 0:
                max_users = user_install_count
                install_source = "User Input"
            else:
                level1_users = df1[df1['LEVEL_CLEAN'] == 1]['USERS'].values[0] if 1 in df1['LEVEL_CLEAN'].values else 0
                level2_users = df1[df1['LEVEL_CLEAN'] == 2]['USERS'].values[0] if 2 in df1['LEVEL_CLEAN'].values else 0
                max_users = max(level1_users, level2_users)
                install_source = f"Auto-calculated (max of Level 1: {level1_users:,}, Level 2: {level2_users:,})"
            
            with st.expander("ℹ️ Show Install Base Info"):
                st.info(f"📊 Using install base of **{max_users:,}** (Source: {install_source})")
            
            df1['Retention %'] = round((df1['USERS'] / max_users) * 100, 2)
            df1['Drop'] = ((df1['USERS'] - df1['USERS'].shift(-1)) / df1['USERS']).fillna(0) * 100
            df1['Drop'] = df1['Drop'].round(2)
            retention_20 = round(df1[df1['LEVEL_CLEAN'] == 20]['Retention %'].values[0], 2) if 20 in df1['LEVEL_CLEAN'].values else 0
            retention_50 = round(df1[df1['LEVEL_CLEAN'] == 50]['Retention %'].values[0], 2) if 50 in df1['LEVEL_CLEAN'].values else 0
            retention_75 = round(df1[df1['LEVEL_CLEAN'] == 75]['Retention %'].values[0], 2) if 75 in df1['LEVEL_CLEAN'].values else 0
            retention_100 = round(df1[df1['LEVEL_CLEAN'] == 100]['Retention %'].values[0], 2) if 100 in df1['LEVEL_CLEAN'].values else 0
            retention_150 = round(df1[df1['LEVEL_CLEAN'] == 150]['Retention %'].values[0], 2) if 150 in df1['LEVEL_CLEAN'].values else 0
            retention_200 = round(df1[df1['LEVEL_CLEAN'] == 200]['Retention %'].values[0], 2) if 200 in df1['LEVEL_CLEAN'].values else 0
        else:
            st.error("❌ Required columns `LEVEL` and `USERS` not found in the Retention Base File.")
            return

        df2 = pd.read_csv(file2)
        df2.columns = df2.columns.str.strip()
        if 'EVENT' in df2.columns and 'USERS' in df2.columns:
            df2 = df2[['EVENT', 'USERS']]
            df2['EVENT_CLEAN'] = df2['EVENT'].apply(lambda x: int(re.search(r"_(\d+)", str(x)).group(1)) if re.search(r"_(\d+)", str(x)) else None)
            df2.dropna(inplace=True)
            df2['EVENT_CLEAN'] = df2['EVENT_CLEAN'].astype(int)
            df2 = pd.concat([pd.DataFrame({'EVENT': ['Assumed_0'], 'USERS': [max_users], 'EVENT_CLEAN': [0]}), df2], ignore_index=True).sort_values('EVENT_CLEAN').reset_index(drop=True)
            df2['% of Users at Ad'] = round((df2['USERS'] / max_users) * 100, 2)
            ad10 = df2[df2['EVENT_CLEAN'] == 10]['% of Users at Ad'].values[0] if 10 in df2['EVENT_CLEAN'].values else 0
            ad20 = df2[df2['EVENT_CLEAN'] == 20]['% of Users at Ad'].values[0] if 20 in df2['EVENT_CLEAN'].values else 0
            ad40 = df2[df2['EVENT_CLEAN'] == 40]['% of Users at Ad'].values[0] if 40 in df2['EVENT_CLEAN'].values else 0
            ad70 = df2[df2['EVENT_CLEAN'] == 70]['% of Users at Ad'].values[0] if 70 in df2['EVENT_CLEAN'].values else 0
            ad100 = df2[df2['EVENT_CLEAN'] == 100]['% of Users at Ad'].values[0] if 100 in df2['EVENT_CLEAN'].values else 0
            df2['Diff of Ads'] = df2['EVENT_CLEAN'].diff().fillna(df2['EVENT_CLEAN']).astype(int)
            df2['Multi1'] = df2['USERS'] * df2['Diff of Ads']
            sum1 = df2['Multi1'].sum()
            df2['Avg Diff Ads'] = df2['Diff of Ads'] / 2
            df2['Diff of Users'] = df2['USERS'].shift(1) - df2['USERS']
            df2['Diff of Users'] = df2['Diff of Users'].fillna(0).astype(int)
            df2['Multi2'] = df2['Avg Diff Ads'] * df2['Diff of Users']
            sum2 = df2['Multi2'].sum()
            avg_ads_per_user = round((sum1 + sum2) / max_users, 2)
            st.success(f"✅ Ad data processed successfully! Total Average Ads per User: **{avg_ads_per_user}**")
        else:
            st.error("❌ Required columns `EVENT` and `USERS` not found in the Ad Event File.")
            return

        st.markdown("---")
        st.subheader("📈 Visualization")
        retention_fig = create_retention_chart(df1[df1['LEVEL_CLEAN'] <= 100], version, date_selected, "Retention Chart (Levels 1-100)")
        drop_fig = create_drop_chart(df1[df1['LEVEL_CLEAN'] <= 100].rename(columns={'LEVEL_CLEAN': 'Level'}), version, date_selected, "Drop Chart (Levels 1-100)")
        st.pyplot(retention_fig)
        st.pyplot(drop_fig)

        st.markdown("---")
        st.subheader("📝 Manual Metrics & Download")
        default_summary_data = {
            "Version": version, "Date Selected": date_selected.strftime("%d-%b-%y"),
            "Check Date": check_date.strftime("%d-%b-%y"), "Install Base": int(max_users),
            "Install Source": install_source, "Total Level Retention (20)": f"{retention_20}%",
            "Total Level Retention (50)": f"{retention_50}%", "Total Level Retention (75)": f"{retention_75}%",
            "Total Level Retention (100)": f"{retention_100}%", "Total Level Retention (150)": f"{retention_150}%",
            "Total Level Retention (200)": f"{retention_200}%", "% of Users at Ad 10": f"{ad10}%",
            "% of Users at Ad 20": f"{ad20}%", "% of Users at Ad 40": f"{ad40}%",
            "% of Users at Ad 70": f"{ad70}%", "% of Users at Ad 100": f"{ad100}%",
            "Avg Ads per User": avg_ads_per_user
        }
        df_summary = pd.DataFrame(list(default_summary_data.items()), columns=["Metric", "Value"])
        
        tab1, tab2 = st.tabs(["📥 Manual Input", "📋 Summary Table"])
        with tab1:
            st.markdown("### 🔧 Enter Manual Metrics Here:")
            day1_retention = st.text_input("Day 1 Retention (%)", value="29.56%")
            day3_retention = st.text_input("Day 3 Retention (%)", value="13.26%")
            session_length = st.text_input("Session Length (in sec)", value="264.5")
            playtime_length = st.text_input("Playtime Length (in sec)", value="936.6")
            if st.button("Update Summary Table"):
                df_summary = df_summary.set_index("Metric")
                df_summary.loc["Day 1 Retention"] = day1_retention
                df_summary.loc["Day 3 Retention"] = day3_retention
                df_summary.loc["Session Length"] = f"{session_length} s"
                df_summary.loc["Playtime Length"] = f"{playtime_length} s"
                df_summary = df_summary.reset_index()
        with tab2:
            st.dataframe(df_summary, hide_index=True, use_container_width=True)

        df_progression = df1[['LEVEL_CLEAN', 'USERS', 'Retention %', 'Drop']].rename(columns={'LEVEL_CLEAN': 'Level'})
        
        excel_data = generate_excel_report_dp1(df_summary, df_progression, retention_fig, drop_fig)
        st.download_button(
            label="📥 Download Excel Report",
            data=excel_data,
            file_name=f"DP1_METRIX_Report_{version}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

# --- Tool 2: GAME PROGRESSION ---
def game_progression_tool():
    st.header("📊 GAME PROGRESSION Dashboard")
    st.markdown("This tool analyzes level start and complete data to track game progression and user drop-offs.")
    
    col1, col2 = st.columns(2)
    with col1:
        start_file = st.file_uploader("📂 Upload Start Level File", type=["xlsx", "csv"], key='start_file')
    with col2:
        complete_file = st.file_uploader("📂 Upload Complete Level File", type=["xlsx", "csv"], key='complete_file')

    if not start_file and not complete_file:
        st.subheader("💡 Sample Input File Format")
        st.markdown("Please use the following format for your files:")
        st.markdown("---")
        st.markdown("#### **Start Level File**")
        st.image("https://i.imgur.com/eBw6j51.png", caption="Sample Start Level Data", use_column_width=True)
        st.markdown("---")
        st.markdown("#### **Complete Level File**")
        st.image("https://i.imgur.com/83u6U54.png", caption="Sample Complete Level Data", use_column_width=True)
        st.markdown("---")

    st.subheader("📌 General Information")
    col3, col4 = st.columns(2)
    with col3:
        version = st.text_input("Game Version", value="1.0.0", key='version2')
    with col4:
        date_selected = st.date_input("Select Date", value=datetime.date.today(), key='date2')

    if start_file and complete_file:
        st.markdown("---")
        st.subheader("✅ Data Processing & Analytics")
        try:
            df_start = pd.read_excel(start_file) if start_file.name.endswith(".xlsx") else pd.read_csv(start_file)
            df_complete = pd.read_excel(complete_file) if complete_file.name.endswith(".xlsx") else pd.read_csv(complete_file)
        except Exception as e:
            st.error(f"❌ Error reading files: {e}")
            return

        df_start.columns = df_start.columns.str.strip().str.upper()
        df_complete.columns = df_complete.columns.str.strip().str.upper()

        level_cols = ['LEVEL', 'LEVELPLAYED', 'TOTALLEVELSPLAYED', 'LEVEL_NUMBER', 'TOTAL_LEVEL']
        level_col_start = next((col for col in df_start.columns if col in level_cols), None)
        user_col_start = next((col for col in df_start.columns if 'USER' in col), None)

        if level_col_start and user_col_start:
            df_start = df_start[[level_col_start, user_col_start]]
            df_start['LEVEL_CLEAN'] = df_start[level_col_start].apply(clean_level)
            df_start.dropna(inplace=True)
            df_start['LEVEL_CLEAN'] = df_start['LEVEL_CLEAN'].astype(int)
            df_start.rename(columns={user_col_start: 'Start Users'}, inplace=True)
        else:
            st.error("❌ Required columns not found in start file.")
            return

        level_col_complete = next((col for col in df_complete.columns if col in level_cols), None)
        user_col_complete = next((col for col in df_complete.columns if 'USER' in col), None)
        additional_cols = [c for c in ['PLAYTIME_AVG', 'PLAY_TIME_AVG', 'HINT_USED_SUM', 'RETRY_COUNT_SUM', 'SKIPPED_SUM', 'ATTEMPT_SUM', 'PREFAB_NAME'] if c in df_complete.columns]

        if level_col_complete and user_col_complete:
            cols_to_keep = [level_col_complete, user_col_complete] + additional_cols
            df_complete = df_complete[cols_to_keep]
            df_complete['LEVEL_CLEAN'] = df_complete[level_col_complete].apply(clean_level)
            df_complete.dropna(inplace=True)
            df_complete['LEVEL_CLEAN'] = df_complete['LEVEL_CLEAN'].astype(int)
            df_complete.rename(columns={user_col_complete: 'Complete Users'}, inplace=True)
        else:
            st.error("❌ Required columns not found in complete file.")
            return

        df = pd.merge(df_start, df_complete, on='LEVEL_CLEAN', how='outer').sort_values('LEVEL_CLEAN')
        base_users = df[df['LEVEL_CLEAN'].isin([1, 2])]['Start Users'].max()
        df['Game Play Drop'] = ((df['Start Users'] - df['Complete Users']) / df['Start Users']) * 100
        df['Popup Drop'] = ((df['Complete Users'] - df['Start Users'].shift(-1)) / df['Complete Users']) * 100
        df['Total Level Drop'] = df['Game Play Drop'] + df['Popup Drop']
        df['Retention %'] = (df['Start Users'] / base_users) * 100

        if 'RETRY_COUNT_SUM' in df.columns:
            df['Attempt'] = df['RETRY_COUNT_SUM'] / df['Complete Users']

        metric_cols = ['Game Play Drop', 'Popup Drop', 'Total Level Drop', 'Retention %']
        if 'Attempt' in df.columns: metric_cols.append('Attempt')
        df[metric_cols] = df[metric_cols].round(2)

        df_100 = df[df['LEVEL_CLEAN'] <= 100]

        st.subheader("📈 Visualization")
        retention_fig = create_retention_chart(df_100, version, date_selected, "Retention Chart (Levels 1-100)")
        drop_fig = create_drop_chart(df_100.rename(columns={'LEVEL_CLEAN': 'Level', 'Total Level Drop': 'Drop'}), version, date_selected, "Total Drop Chart (Levels 1-100)")
        st.pyplot(retention_fig)
        st.pyplot(drop_fig)

        st.markdown("---")
        st.subheader("📉 Combo Drop Chart (Levels 1-100)")
        drop_comb_fig, ax3 = plt.subplots(figsize=(15, 6))
        width = 0.4
        x = df_100['LEVEL_CLEAN']
        ax3.bar(x + width/2, df_100['Game Play Drop'], width, color='#66BB6A', label='Game Play Drop')
        ax3.bar(x - width/2, df_100['Popup Drop'], width, color='#42A5F5', label='Popup Drop')
        ax3.set_xlim(1, 100)
        max_drop = max(df_100['Game Play Drop'].max(), df_100['Popup Drop'].max())
        ax3.set_ylim(0, max(max_drop, 10) + 10)
        ax3.set_xticks(np.arange(1, 101, 1))
        xtick_labels = [f"$\\bf{{{val}}}$" if val % 5 == 0 else str(val) for val in np.arange(1, 101, 1)]
        ax3.set_xticklabels(xtick_labels, fontsize=6)
        ax3.set_yticks(np.arange(0, max(max_drop, 10) + 11, 5))
        ax3.set_xlabel("Level")
        ax3.set_ylabel("% Of Users Dropped")
        ax3.set_title(f"Game Play & Popup Drop Chart | Version {version} | Date: {date_selected.strftime('%d-%m-%Y')}", fontsize=12, fontweight='bold')
        ax3.grid(True, linestyle='--', linewidth=0.5)
        ax3.legend(loc='upper right', fontsize=8)
        plt.tight_layout()
        st.pyplot(drop_comb_fig)

        st.markdown("---")
        st.subheader("⬇️ Download Excel Report")
        export_cols = ['LEVEL_CLEAN', 'Start Users', 'Complete Users', 'Game Play Drop', 'Popup Drop', 'Total Level Drop', 'Retention %'] + additional_cols
        df_export = df[export_cols].rename(columns={'LEVEL_CLEAN': 'Level'})
        st.dataframe(df_export, hide_index=True, use_container_width=True)
        
        excel_data_gp = generate_excel_report_gp(df_export, retention_fig, drop_fig, drop_comb_fig, version, date_selected)
        st.download_button(
            label="📥 Download Excel Report",
            data=excel_data_gp,
            file_name=f"Game_Progression_Report_{version}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )


# --- Tool 3: GAME LEVEL DATA ANALYZER ---
def process_files_analyzer(start_df, complete_df):
    def get_column(df, possible_names):
        for col in df.columns:
            if col.strip().lower() in [name.lower() for name in possible_names]:
                return col
        return None
    
    start_df.columns = start_df.columns.str.strip().str.upper()
    complete_df.columns = complete_df.columns.str.strip().str.upper()

    level_col = get_column(start_df, ['LEVEL', 'TOTALLEVELS', 'STAGE'])
    game_col = get_column(start_df, ['GAME_ID', 'CATEGORY', 'Game_name' , 'MISSION'])
    diff_col = get_column(start_df, ['DIFFICULTY', 'mode'])
    
    playtime_col = get_column(complete_df, ['PLAY_TIME_AVG', 'PLAYTIME', 'PLAYTIME_AVG', 'playtime_avg'])
    hint_col = get_column(complete_df, ['HINT_USED_SUM', 'HINT_USED', 'HINT'])
    skipped_col = get_column(complete_df, ['SKIPPED_SUM', 'SKIPPED', 'SKIP'])
    attempts_col = get_column(complete_df, ['ATTEMPTS_SUM', 'ATTEMPTS', 'TRY_COUNT'])
    
    for df in [start_df, complete_df]:
        if level_col:
            df[level_col] = df[level_col].apply(clean_level)
            df.sort_values(level_col, inplace=True)
            
    rename_dict_start = {'USERS': 'Start Users'}
    if level_col: rename_dict_start[level_col] = 'LEVEL'
    if game_col: rename_dict_start[game_col] = 'GAME_ID'
    if diff_col: rename_dict_start[diff_col] = 'DIFFICULTY'
    start_df.rename(columns=rename_dict_start, inplace=True)

    rename_dict_complete = {'USERS': 'Complete Users'}
    if level_col: rename_dict_complete[level_col] = 'LEVEL'
    if game_col: rename_dict_complete[game_col] = 'GAME_ID'
    if diff_col: rename_dict_complete[diff_col] = 'DIFFICULTY'
    if playtime_col: rename_dict_complete[playtime_col] = 'PLAY_TIME_AVG'
    if hint_col: rename_dict_complete[hint_col] = 'HINT_USED_SUM'
    if skipped_col: rename_dict_complete[skipped_col] = 'SKIPPED_SUM'
    if attempts_col: rename_dict_complete[attempts_col] = 'ATTEMPTS_SUM'
    complete_df.rename(columns=rename_dict_complete, inplace=True)

    merge_cols = [col for col in ['GAME_ID', 'DIFFICULTY', 'LEVEL'] if col in start_df.columns]
    merged = pd.merge(start_df, complete_df, on=merge_cols, how='outer')
    
    keep_cols = merge_cols + ['Start Users', 'Complete Users']
    if 'PLAY_TIME_AVG' in merged.columns: keep_cols.append('PLAY_TIME_AVG')
    if 'HINT_USED_SUM' in merged.columns: keep_cols.append('HINT_USED_SUM')
    if 'SKIPPED_SUM' in merged.columns: keep_cols.append('SKIPPED_SUM')
    if 'ATTEMPTS_SUM' in merged.columns: keep_cols.append('ATTEMPTS_SUM')
    merged = merged[[col for col in keep_cols if col in merged.columns]]

    if 'Start Users' in merged.columns and 'Complete Users' in merged.columns:
        merged['Game Play Drop'] = ((merged['Start Users'] - merged['Complete Users']) / merged['Start Users'].replace(0, np.nan)) * 100
        merged['Popup Drop'] = ((merged['Complete Users'] - merged['Start Users'].shift(-1)) / merged['Complete Users'].replace(0, np.nan)) * 100
    else:
        merged['Game Play Drop'] = 0
        merged['Popup Drop'] = 0

    def calculate_retention(group):
        if 'Start Users' not in group.columns:
            group['Retention %'] = 0
            return group
        
        base_users = group[group['LEVEL'].isin([1, 2])]['Start Users'].max()
        if base_users == 0 or pd.isnull(base_users):
            base_users = group['Start Users'].max()
        
        group['Retention %'] = (group['Start Users'] / base_users) * 100
        return group

    group_cols = [col for col in ['GAME_ID', 'DIFFICULTY'] if col in merged.columns]
    if not group_cols:
        merged['All Data'] = 'All Data'
        group_cols = ['All Data']
        
    merged = merged.groupby(group_cols, group_keys=False).apply(calculate_retention)
    
    fill_cols = ['Start Users', 'Complete Users', 'PLAY_TIME_AVG', 'HINT_USED_SUM', 'SKIPPED_SUM', 'ATTEMPTS_SUM']
    merged.fillna({col: 0 for col in fill_cols if col in merged.columns}, inplace=True)
    
    if 'Game Play Drop' in merged.columns and 'Popup Drop' in merged.columns:
        merged['Total Level Drop'] = merged['Game Play Drop'] + merged['Popup Drop']
    else:
        merged['Total Level Drop'] = 0
        
    return merged

def create_charts_analyzer(df, game_name):
    charts = {}
    df_100 = df[df['LEVEL'] <= 100]
    xtick_labels = [f"$\\bf{{{val}}}$" if val % 5 == 0 else str(val) for val in np.arange(1, 101, 1)]
    
    fig1, ax1 = plt.subplots(figsize=(15, 5))
    if 'Retention %' in df_100.columns and not df_100['Retention %'].dropna().empty:
        ax1.plot(df_100['LEVEL'], df_100['Retention %'], linestyle='-', color='#F57C00', linewidth=2, label='Retention')
        ax1.set_xlim(1, 100)
        ax1.set_ylim(0, 110)
        ax1.set_xticks(np.arange(1, 101, 1))
        ax1.set_yticks(np.arange(0, 111, 5))
        ax1.set_xticklabels(xtick_labels, fontsize=4)
        ax1.set_title(f"{game_name} | Retention Chart (Levels 1–100)", fontsize=12, fontweight='bold')
        ax1.grid(True, linestyle='--', linewidth=0.5)
        ax1.legend(loc='lower left', fontsize=8)
        charts['retention'] = fig1
        
    fig2, ax2 = plt.subplots(figsize=(15, 5))
    if 'Total Level Drop' in df_100.columns and not df_100['Total Level Drop'].dropna().empty:
        ax2.bar(df_100['LEVEL'], df_100['Total Level Drop'], color='#EF5350', label='Drop Rate')
        drop_max = df_100['Total Level Drop'].max() if not df_100['Total Level Drop'].dropna().empty else 0
        ymax = max(drop_max, 10) + 10
        ax2.set_xlim(1, 100)
        ax2.set_ylim(0, ymax)
        ax2.set_xticks(np.arange(1, 101, 1))
        ax2.set_yticks(np.arange(0, ymax + 1, 5))
        ax2.set_xticklabels(xtick_labels, fontsize=4)
        ax2.set_title(f"{game_name} | Total Drop Chart (Levels 1–100)", fontsize=12, fontweight='bold')
        ax2.grid(True, linestyle='--', linewidth=0.5)
        ax2.legend(loc='upper right', fontsize=8)
        charts['total_drop'] = fig2

    fig3, ax3 = plt.subplots(figsize=(15, 5))
    if 'Game Play Drop' in df_100.columns and 'Popup Drop' in df_100.columns and not df_100['Game Play Drop'].dropna().empty:
        width = 0.4
        x = df_100['LEVEL']
        ax3.bar(x - width/2, df_100['Popup Drop'], width, color='#42A5F5', label='Popup Drop')
        ax3.bar(x + width/2, df_100['Game Play Drop'], width, color='#66BB6A', label='Game Play Drop')
        ax3.set_xlim(1, 100)
        ax3.set_xticks(np.arange(1, 101, 1))
        ax3.set_xticklabels(xtick_labels, fontsize=4)
        ax3.set_title(f"{game_name} | Game Play & Popup Drop (Levels 1–100)", fontsize=10, fontweight='bold')
        ax3.legend(loc='upper right', fontsize=6)
        ax3.grid(True, linestyle='--', linewidth=0.5)
        charts['combined_drop'] = fig3
        
    return charts

def generate_excel_analyzer(processed_data):
    wb = Workbook()
    wb.remove(wb.active)
    
    main_sheet = wb.create_sheet("MAIN_TAB")
    main_headers = ["Index", "Sheet Name", "Link to Sheet", "LEVEL_Start", "Start Users",
                    "LEVEL_End", "USERS_END", "Game Play Drop Count", "Popup Drop Count",
                    "Total Level Drop Count"]
    main_sheet.append(main_headers)
    
    for col in main_sheet[1]:
        col.font = Font(bold=True, color="FFFFFF")
        col.fill = PatternFill("solid", fgColor="4F81BD")
    
    fill_dark_red = PatternFill("solid", fgColor="8B0000")
    font_white = Font(color="FFFFFF")
    fill_red = PatternFill("solid", fgColor="CD5C5C")
    fill_pink = PatternFill("solid", fgColor="FFC0CB")
    font_black = Font(color="000000")

    for idx, (game_key, df) in enumerate(processed_data.items(), start=1):
        sheet_name = str(game_key)[:31]
        ws = wb.create_sheet(sheet_name)
        
        headers = ["=HYPERLINK(\"#MAIN_TAB!A1\", \"Back to Main\")"] + [col for col in df.columns if col not in ['GAME_ID', 'DIFFICULTY']]
        ws.append(headers)
        
        ws['A1'].font = Font(color="0000FF", underline="single", bold=True, size=14)
        ws['A1'].fill = PatternFill("solid", fgColor="FFFF00")
        ws.column_dimensions['A'].width = 25
        
        for r_idx, row in df.iterrows():
            row_values = [
                row.get('LEVEL', 0), row.get('Start Users', 0), row.get('Complete Users', 0),
                round(row.get('Game Play Drop', 0), 2), round(row.get('Popup Drop', 0), 2),
                round(row.get('Total Level Drop', 0), 2), round(row.get('Retention %', 0), 2),
                round(row.get('PLAY_TIME_AVG', 0), 2), round(row.get('HINT_USED_SUM', 0), 2),
                round(row.get('SKIPPED_SUM', 0), 2), round(row.get('ATTEMPTS_SUM', 0), 2)
            ]
            
            ws.append(row_values)
            
            # Apply color formatting
            start_col_idx = 4
            for i, col_name in enumerate(['Game Play Drop', 'Popup Drop', 'Total Level Drop']):
                if col_name in df.columns:
                    value = row.get(col_name, 0)
                    cell = ws.cell(row=ws.max_row, column=start_col_idx + i)
                    if value >= 10:
                        cell.fill = fill_dark_red
                        cell.font = font_white
                    elif value >= 5:
                        cell.fill = fill_red
                        cell.font = font_white
                    elif value >= 3:
                        cell.fill = fill_pink
                        cell.font = font_black
                        
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output

def game_level_analyzer_tool():
    st.header("📊 GAME LEVEL DATA ANALYZER")
    st.markdown("This tool processes level start and complete data, providing detailed analytics per game and difficulty.")
    
    col1, col2 = st.columns(2)
    with col1:
        start_file = st.file_uploader("📂 Upload Start Level File", type=["xlsx", "csv"], key='analyzer_start')
    with col2:
        complete_file = st.file_uploader("📂 Upload Complete Level File", type=["xlsx", "csv"], key='analyzer_complete')

    if not start_file and not complete_file:
        st.subheader("💡 Sample Input File Format")
        st.markdown("Please ensure your data includes columns for `level`, `users`, and optionally `game_id` and `difficulty`.")
        st.image("https://i.imgur.com/83u6U54.png", caption="Sample Data for Analyzer Tool", use_column_width=True)

    if start_file and complete_file:
        st.markdown("---")
        st.subheader("✅ Data Processing & Analytics")
        try:
            start_df = pd.read_excel(start_file) if start_file.name.endswith(".xlsx") else pd.read_csv(start_file)
            complete_df = pd.read_excel(complete_file) if complete_file.name.endswith(".xlsx") else pd.read_csv(complete_file)
        except Exception as e:
            st.error(f"❌ Error reading files: {e}")
            return
            
        processed_data = process_files_analyzer(start_df, complete_df)
        
        unique_games = processed_data.groupby(['GAME_ID', 'DIFFICULTY']).groups.keys()
        game_options = [f"{game} ({diff})" if 'DIFFICULTY' in processed_data.columns else str(game) for game, diff in unique_games]
        game_options.insert(0, "All Data")
        
        selected_option = st.selectbox("🎯 Select Game & Difficulty to Analyze", options=game_options)
        
        if selected_option == "All Data":
            filtered_df = processed_data
            charts = create_charts_analyzer(filtered_df, "All Data")
        else:
            game, diff = selected_option.replace("(", "").replace(")", "").split(" (")
            filtered_df = processed_data[(processed_data['GAME_ID'] == game) & (processed_data['DIFFICULTY'] == diff)]
            charts = create_charts_analyzer(filtered_df, selected_option)
            
        st.subheader("📊 Analyzed Data Table")
        st.dataframe(filtered_df, use_container_width=True, hide_index=True)
        
        st.subheader("📈 Visualizations")
        if 'retention' in charts:
            st.pyplot(charts['retention'])
        if 'total_drop' in charts:
            st.pyplot(charts['total_drop'])
        if 'combined_drop' in charts:
            st.pyplot(charts['combined_drop'])

        st.subheader("⬇️ Download Detailed Excel Report")
        with st.spinner("Generating Excel report..."):
            excel_data = generate_excel_analyzer(processed_data.groupby(['GAME_ID', 'DIFFICULTY'] if 'DIFFICULTY' in processed_data.columns else 'All Data'))
            st.download_button(
                label="📥 Download Excel Report",
                data=excel_data,
                file_name=f"Game_Level_Analyzer_Report.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        

# --- Main App Logic ---
if __name__ == "__main__":
    if check_password():
        st.set_page_config(page_title="Game Analytics Tools", layout="wide")
        st.title("🎮 Game Analytics Dashboard")
        
        tab_list = ["GAME PROGRESSION", "GAME LEVEL DATA ANALYZER"]
        selected_tool = st.sidebar.radio("Select a Tool", tab_list)

        if selected_tool == "GAME PROGRESSION":
            game_progression_tool()
        elif selected_tool == "GAME LEVEL DATA ANALYZER":
            game_level_analyzer_tool()
